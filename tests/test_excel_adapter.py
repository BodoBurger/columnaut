from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from columnaut.ingestion.excel_adapter import ExcelAdapter
from columnaut.models import LoadOptions


def workbook_bytes() -> bytes:
    buffer = BytesIO()
    raw = pd.DataFrame(
        [
            ["Account", "Amount", "Amount", None, "Status"],
            [1, 10, 20, None, "open"],
            [2, "oops", 30, None, "closed"],
            [None, None, None, None, None],
            [3, 40, 50, None, "open"],
        ]
    )
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        raw.to_excel(writer, sheet_name="Transactions", header=False, index=False)
        pd.DataFrame({"code": ["DE", "FR"]}).to_excel(
            writer, sheet_name="Countries", index=False
        )

    editable = load_workbook(BytesIO(buffer.getvalue()))
    editable["Transactions"].merge_cells("F1:G1")
    editable["Transactions"]["F1"] = "Merged note"
    output = BytesIO()
    editable.save(output)
    editable.close()
    return output.getvalue()


def test_excel_inspection_lists_sheets() -> None:
    inspection = ExcelAdapter().inspect(workbook_bytes(), "sample.xlsx")

    assert inspection.sheet_names == ("Transactions", "Countries")


def test_excel_load_honors_sheet_and_header_row() -> None:
    loaded = ExcelAdapter().load(
        workbook_bytes(),
        "sample.xlsx",
        LoadOptions(sheet_name="Countries", header_row=0),
    )

    assert list(loaded.dataframe.columns) == ["code"]
    assert loaded.dataframe["code"].tolist() == ["DE", "FR"]
    assert isinstance(loaded.dataframe["code"].dtype, pd.ArrowDtype)


def test_excel_load_reports_structural_problems() -> None:
    loaded = ExcelAdapter().load(
        workbook_bytes(),
        "sample.xlsx",
        LoadOptions(sheet_name="Transactions", header_row=0),
    )

    warning_codes = {warning.code for warning in loaded.warnings}
    assert loaded.dataframe.columns[:5].tolist() == [
        "Account",
        "Amount",
        "Amount__2",
        "column_4",
        "Status",
    ]
    assert {
        "duplicate_headers",
        "header_gaps",
        "merged_cells",
        "empty_rows",
        "empty_columns",
        "mixed_value_types",
    } <= warning_codes
    assert isinstance(loaded.dataframe["Account"].dtype, pd.ArrowDtype)
    assert loaded.dataframe["Amount"].dtype == object


def test_excel_adapter_declares_legacy_xls_support() -> None:
    assert ExcelAdapter().supports("legacy.xls")
